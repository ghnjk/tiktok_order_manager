#!/usr/bin/env python3
# -*- coding:utf-8 _*-
"""
@file: tiktok_order
@author: jkguo
@create: 2023/1/15
"""
import json
import logging
import os.path
import shutil
import typing
import time
from json_serialize import JsonSerializable

from openpyxl import load_workbook

SKU_MAP: typing.Dict[str, typing.List[str]] = {
}


class SkuMapper(object):

    def __init__(self):
        self.order_sku_info: typing.Dict[str, typing.List[dict]] = {}
        self.order_cod_payments: typing.Dict[str, float] = {}

    def load_sku_map(self, sku_order_xls: str):
        wb = load_workbook(filename=sku_order_xls)
        # 打开第一个sheet
        ws = wb[wb.sheetnames[0]]
        row_idx = 0
        while True:
            row_idx += 1
            order_no = ws.cell(row=row_idx, column=1).value
            if order_no is None:
                break
            sku = ws.cell(row=row_idx, column=2).value
            count = str(ws.cell(row=row_idx, column=3).value)
            if not count.isdigit():
                continue
            count = int(count)
            item_price = float(ws.cell(row=row_idx, column=4).value)
            if order_no not in self.order_sku_info:
                self.order_sku_info[order_no] = []
            self.order_sku_info[order_no].append({
                "sku": sku,
                "count": count,
                "item_price": item_price
            })
            cod_amount = str(ws.cell(row=row_idx, column=6).value)
            cod_amount = float(cod_amount.replace(",", ""))
            self.order_cod_payments[order_no] = cod_amount

    def get_sku_list(self, tiktok_order_id: str):
        return self.order_sku_info.get(tiktok_order_id, None)

    def get_cod_amount(self, tiktok_order_id: str):
        return self.order_cod_payments.get(tiktok_order_id, None)


class TikTokOrder(JsonSerializable):

    def __init__(self):
        self.track_order: str = ""
        self.sender_addr: str = ""
        self.receiver_name: str = ""
        self.receiver_phone: str = ""
        self.receiver_addr: str = ""
        self.weight: str = ""
        self.goods: str = ""
        self.sku_list: typing.List[dict] = []
        self.payment: str = "COD"
        self.tiktok_order_id: str = ""
        self.cod: str = ""
        self.sender_name: str = ""
        self.sender_phone: str = ""
        self.price: float = 0
        self.pdf_file_path: str = ""

    def __str__(self):
        return json.dumps(self.to_dict(), indent=4)

    def parse(self, text_list: typing.List[str], sku_mapper: SkuMapper):
        # for line in text_list:
        #     print("======")
        #     print(line)
        for i in range(len(text_list)):
            if i == 1:
                self.track_order = text_list[i]
            elif i == 3:
                self.sender_name = text_list[i]
            elif i == 4:
                self.sender_addr = text_list[i]
            elif i in [5, 6]:
                if text_list[i].strip().startswith("Receiver:"):
                    self.receiver_addr = text_list[i]
            if i in (7, 8):
                if text_list[i].strip().startswith("Goods:"):
                    self.goods = text_list[i].split("\n")[0][7:]
                elif text_list[i].startswith("Weigh"):
                    self.weight = text_list[i]
                else:
                    s = text_list[i].strip()
                    if s.find("+63") >= 0:
                        self.receiver_phone = s
            if i >= 7:
                # cod收费和tiktok_order_id可能顺序是乱的
                if len(text_list[i].strip()) == 18 and text_list[i].strip().isdigit():
                    self.tiktok_order_id = text_list[i].strip()
                else:
                    s = text_list[i].replace("\n", " ").strip()
                    if s.startswith("COD :"):
                        self.cod = s[5:].replace("PHP", "").strip()
        if len(self.tiktok_order_id) == 0:
            raise Exception("parse failed. tiktok_order_id is empty")
        self.sku_list = sku_mapper.get_sku_list(self.tiktok_order_id)
        if self.sku_list is None:
            raise Exception(f"tiktok_order_id {self.tiktok_order_id} not find sku xlsx")
        # format price
        self.format_price()
        cod_price = sku_mapper.get_cod_amount(self.tiktok_order_id)
        if cod_price is None or (cod_price != self.price and self.cod != "0"):
            raise Exception(
                f"tiktok_order_id {self.tiktok_order_id} cod amount {self.price} not consist with {cod_price}")
        # 格式化接收人信息
        self.format_receiver_info()
        # 格式化发送人信息
        self.format_sender_info()
        # 重新设置卖家信息
        self.rewrite_sender_info()

    def to_xls_row(self) -> typing.List[typing.List[str]]:
        is_free = False
        if self.cod == "0" and self.price < 1e-4:
            is_free = True
        rows = []
        for sku_info in self.sku_list:
            sku = sku_info["sku"]
            count = sku_info["count"]
            item_price = sku_info["item_price"]
            if is_free:
                item_price = 0
            rows.append(
                [
                    self.tiktok_order_id,
                    self.track_order,
                    sku,
                    str(count),
                    "{:.2f}".format(item_price),
                    "",
                    self.receiver_name,
                    self.receiver_phone,
                    self.receiver_addr,
                    "",
                    "",
                    self.payment,
                    "",
                    "",
                    "J&T EXPRESS",
                    "",
                    "",
                    self.sender_name,
                    self.sender_phone,
                    self.sender_addr,
                    "",
                    ""
                ]
            )
        return rows

    def is_valid_order(self):
        if len(self.tiktok_order_id) == 0:
            logging.error("tiktok_order_id is empty")
            return False, "tiktok_order_id is empty"
        if len(self.track_order) == 0:
            logging.error("track_order is empty")
            return False, "track_order is empty"
        if self.sku_list is None or len(self.sku_list) == 0:
            logging.error("sku_list is empty")
            return False, "sku_list is empty"
        if self.price < 0:
            logging.error(f"price {self.price} invalid")
            return False, f"price {self.price} invalid"
        if len(self.receiver_name) == 0:
            logging.error(f"receiver_name is empty")
            return False, f"receiver_name is empty"
        if len(self.receiver_phone) == 0:
            logging.error(f"receiver_phone is empty")
            return False, f"receiver_phone is empty"
        if len(self.receiver_addr) == 0:
            logging.error(f"receiver_addr is empty")
            return False, f"receiver_addr is empty"
        if len(self.sender_name) == 0:
            logging.error(f"sender_name is empty")
            return False, f"sender_name is empty"
        if len(self.sender_addr) == 0:
            logging.error(f"sender_addr is empty")
            return False, f"sender_addr is empty"
        return True, ""

    def format_price(self):
        self.cod = self.cod.replace(",", "").strip()
        if self.cod == "0":
            self.price = 0.0
            return
        # 解析价格
        s = self.cod
        ps = ""
        for c in s:
            if c.isdigit() or c == ".":
                ps += c
            else:
                break
        self.price = float(ps)
        if self.price <= 1e-4 and self.cod != "0":
            logging.error(f"invalid price {self.price}")
            raise Exception(f"invalid price {self.price}")
        # 核对sku的价格
        # 价格验证
        sum_p = 0.0
        for sku in self.sku_list:
            sum_p += sku["item_price"] * sku["count"]
        if self.price - sum_p < -1e-4:
            logging.error(f"sum_p {sum_p} bigger than price {self.price} ")
            raise Exception(f"sum_p {sum_p} bigger than price {self.price} ")
        else:
            diff_p = self.price - sum_p
            self.sku_list[0]["item_price"] += diff_p / self.sku_list[0]["count"]
        sum_p = 0.0
        for sku in self.sku_list:
            sum_p += sku["item_price"] * sku["count"]
        if abs(self.price - sum_p) > 1e-4:
            logging.error(f"sum_p {sum_p} not equal to price {self.price} ")
            raise Exception(f"sum_p {sum_p} not equal to price {self.price} ")

    def format_receiver_info(self):
        self.receiver_addr = self.receiver_addr.strip()
        if not self.receiver_addr.startswith("Receiver:"):
            logging.info(f"invalid receiver info {self.receiver_addr}")
            raise Exception(f"invalid receiver info {self.receiver_addr}")
        # 解析名称
        first_line = self.receiver_addr[9:].split("\n")[0].strip()
        if first_line.find("(+63)") >= 0:
            self.receiver_name = first_line[: first_line.find("(+63)")].strip()
        else:
            self.receiver_name = first_line.strip()
        # 解析手机号
        if len(self.receiver_phone) == 0:
            s = self.receiver_addr
            idx = s.find("(+63)")
            if idx >= 0:
                idx += 5
                s = s[idx:].strip()
                for idx in range(min(20, len(s))):
                    if not s[idx].isdigit():
                        break
                self.receiver_phone = s[: idx].strip()
        # 格式化接受人信息
        # 去掉第一行
        idx = self.receiver_addr.find("\n")
        s = self.receiver_addr[idx:].strip()
        if s.startswith("(+63)"):
            i = 0
            for i in range(5, 30):
                if not s[i].isdigit():
                    break
            s = s[i:]
        self.receiver_addr = s.strip()

    def format_sender_info(self):
        s = self.sender_addr.strip()
        if not s.startswith("Sender:"):
            logging.info(f"invalid sender info {self.sender_addr}")
            raise Exception(f"invalid sender info {self.sender_addr}")
        self.sender_addr = s[7:].strip()

    def rewrite_sender_info(self):
        self.sender_name = "Jenny Que"
        self.sender_phone = "09298645333"
        self.sender_addr = "yellow green building, Edano st.,Leon ginto st.,BRGY.9 lucena city(near central lumber)"


def append_orders_to_xls(order_list: typing.List[TikTokOrder], xls_file_path: str,
                         template_xls_file_path: str = "./data/import_hand_order_template_cn.xlsx"):
    wb = load_workbook(filename=template_xls_file_path)
    # 打开第一个sheet
    ws = wb[wb.sheetnames[0]]
    # row_idx = 0
    # while True:
    #     row_idx += 1
    #     v = ws.cell(row=row_idx, column=1).value
    #     if v is None:
    #         break
    #     print(v)
    for order in order_list:
        for r in order.to_xls_row():
            ws.append(r)
        # print(f"add {order.tiktok_order_id}")
    wb.save(xls_file_path)
    wb.close()


def save_order_to_db(pdf_file_path: str, order: TikTokOrder):
    db_dir = "./data/db"
    today = time.strftime("%Y%m%d", time.localtime())
    if not os.path.isdir(db_dir):
        os.mkdir(db_dir)
    db_dir = f"{db_dir}/{today}"
    if not os.path.isdir(db_dir):
        os.mkdir(db_dir)
    backup_pdf_file = os.path.join(db_dir, f"{order.tiktok_order_id}.pdf")
    shutil.copyfile(pdf_file_path, backup_pdf_file)
    order_info_file = os.path.join(db_dir, f"{order.tiktok_order_id}.json")
    with open(order_info_file, "w", encoding="utf-8") as fp:
        json.dump(order.to_dict(), fp, indent=4)
    return backup_pdf_file


def __gen_sku_tmp_pdf(order: TikTokOrder):
    """
    用于覆盖原来面单的sku信息的pdf页面
    pdf 页面大小： (宽*高）10.52 × 14.82厘米

    The pagesize argument is a tuple of two numbers in points (1/72 of an inch).

    :param order:
    :return:
    """
    from reportlab.pdfgen import canvas
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    today = time.strftime("%Y%m%d", time.localtime())
    # create sku tmp pdf
    sku_tmp_pdf_path = f"./tmp/{today}/{order.tiktok_order_id}.sku.pdf"
    if not os.path.isdir(os.path.dirname(sku_tmp_pdf_path)):
        os.mkdir(os.path.dirname(sku_tmp_pdf_path))
    cm_per_inch = 2.54
    # (288, 360)
    c = canvas.Canvas(sku_tmp_pdf_path, pagesize=(
        10.52 / cm_per_inch * 72,
        14.82 / cm_per_inch * 72
    ))
    data = []
    i = 0
    for sku in order.sku_list:
        if i % 5 == 0:
            data.append([])
        data[-1].append(sku["sku"])
        data[-1].append("* {}".format(sku["count"]))
        i += 1
    while i < 10:
        if i % 5 == 0:
            data.append([])
        data[-1].append("-----")
        data[-1].append("   ")
        i += 1
    table = Table(data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("GRID", (1, 0), (1, -1), 1, colors.black),
        ("GRID", (3, 0), (3, -1), 1, colors.black),
        ("GRID", (5, 0), (5, -1), 1, colors.black),
        ("GRID", (7, 0), (7, -1), 1, colors.black),
        ("GRID", (9, 0), (9, -1), 1, colors.black)
    ]))
    table.wrapOn(c, 20, 50)
    table.drawOn(c, 12, 4)
    c.showPage()
    c.save()
    return sku_tmp_pdf_path


def generate_print_pdf(order_list: typing.List[TikTokOrder], print_all_pdf_file: str):
    from PyPDF2 import PdfReader, PdfWriter
    from copy import copy
    with PdfWriter(print_all_pdf_file) as print_all:
        for order in order_list:
            sku_tmp_pdf_path = __gen_sku_tmp_pdf(order)
            sku_reader = PdfReader(sku_tmp_pdf_path)
            r = PdfReader(order.pdf_file_path)
            source_page = r.pages[0]
            gen_page = copy(source_page)
            gen_page.merge_page(sku_reader.pages[0])
            print_all.add_page(gen_page)
